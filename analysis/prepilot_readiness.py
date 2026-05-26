#!/usr/bin/env python3
"""
Build a pre-pilot readiness report for the CSSL Validation task.

The report is intentionally operational: it checks whether the task assets,
deterministic schedules, QA outputs, simulation diagnostics, and optional pilot
workbooks are coherent enough to start or continue pilot testing.
"""

from __future__ import annotations

import argparse
import csv
import json
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUT_DIR = ROOT / "analysis" / "qa_outputs"
STIMULUS_TOOLS = ROOT / "stimulus_tools"
sys.path.insert(0, str(STIMULUS_TOOLS))

from build_plan2_schedule import build_schedule, object_visual_family  # noqa: E402


REQUIRED_FILES = [
    "index.html",
    "styles.css",
    "task.js",
    "config/task_design_plan2.json",
    "stimuli/participant_lists_plan2.json",
    "stimuli/participant_lists_plan2.csv",
    "audio/manifest.json",
]

REQUIRED_WORKBOOK_SHEETS = [
    "Metadata",
    "Summary",
    "Data",
    "ModelReady",
    "LearningEvents",
    "LearningTrials",
    "Practice",
    "PairMap",
    "LearningSchedule",
    "TestSchedule",
    "Config",
    "Notes",
]

REQUIRED_MODEL_READY_COLUMNS = [
    "participantId",
    "listId",
    "observationSeq",
    "observationType",
    "block",
    "pairId",
    "word",
    "targetObjectId",
    "contrast",
    "contrastGroup",
    "phonology",
    "syllableCount",
    "syllableTemplate",
    "phones",
    "ipaTarget",
    "phonologicalNeighborhoodSize",
    "nearestRealWordDistance",
    "nearestRealWords",
    "encountersCompletedForWord",
    "choiceSetSize",
    "chanceLevel",
    "optionObjectIds",
    "optionPairIds",
    "targetPosition",
    "responseObjectId",
    "responsePosition",
    "responseSource",
    "correct",
    "noResponse",
    "timedOut",
    "rtMs",
    "previousResponseObjectId",
    "previousCorrect",
    "previousIncorrect",
    "hasPreviousForWord",
    "previousResponseInChoiceSet",
    "sameResponseAsPrevious",
    "switchedFromPreviousResponse",
    "maintainedAvailablePreviousResponse",
    "switchedAwayFromAvailablePreviousResponse",
    "forcedSwitchFromUnavailablePreviousResponse",
    "audioSource",
    "audioPlayOk",
    "replayAllowed",
    "replayCount",
]

EXPECTED_MAIN_AUDIO = 40
EXPECTED_PRACTICE_AUDIO = 5
EXPECTED_MAIN_IMAGES = 40
EXPECTED_PRACTICE_IMAGES = 5
EXPECTED_WORDS_PER_LIST = 20
EXPECTED_BLOCKS = 5
EXPECTED_LEARNING_TRIALS_PER_BLOCK = 20
EXPECTED_TEST_TRIALS_PER_BLOCK = 20
EXPECTED_LEARNING_EVENTS = 300
EXPECTED_LEARNING_TRIALS = 100
EXPECTED_TEST_TRIALS = 100
EXPECTED_MODEL_READY_ROWS = 400


@dataclass
class CheckRow:
    status: str
    area: str
    check: str
    detail: str


def add_check(rows: list[CheckRow], status: str, area: str, check: str, detail: str) -> None:
    rows.append(CheckRow(status=status, area=area, check=check, detail=detail))


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path: Path, rows: list[CheckRow]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["status", "area", "check", "detail"])
        writer.writeheader()
        for row in rows:
            writer.writerow(row.__dict__)


def parse_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_int(value: Any) -> int | None:
    numeric = parse_float(value)
    return int(numeric) if numeric is not None else None


def truthy(value: Any) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y", "fail", "failed"}


def falsey(value: Any) -> bool:
    return str(value).strip().lower() in {"0", "false", "no", "n"}


def status_rank(status: str) -> int:
    return {"PASS": 0, "INFO": 0, "WARN": 1, "FAIL": 2}.get(status, 0)


def overall_status(rows: list[CheckRow]) -> str:
    if any(row.status == "FAIL" for row in rows):
        return "FAIL"
    if any(row.status == "WARN" for row in rows):
        return "WARN"
    return "PASS"


def run_refresh(args: argparse.Namespace, rows: list[CheckRow]) -> None:
    commands = [
        [sys.executable, str(ROOT / "analysis" / "prepare_audio_qa.py")],
        [
            sys.executable,
            str(ROOT / "analysis" / "qa_image_similarity.py"),
            "--participants",
            str(args.participants),
        ],
    ]
    if not args.skip_simulation:
        commands.extend([
            [
                sys.executable,
                str(ROOT / "analysis" / "run_simulation_scenarios.py"),
                "--participants",
                str(args.participants),
            ],
            [
                sys.executable,
                str(ROOT / "analysis" / "benchmark_analysis_methods.py"),
                "--participants",
                str(args.participants),
            ],
        ])

    for command in commands:
        result = subprocess.run(command, cwd=ROOT, text=True, capture_output=True, check=False)
        name = Path(command[1]).name
        if result.returncode == 0:
            add_check(rows, "PASS", "refresh", name, "completed")
        else:
            detail = (result.stderr or result.stdout or "command failed").strip().splitlines()
            add_check(rows, "FAIL", "refresh", name, detail[-1] if detail else "command failed")


def check_required_files(rows: list[CheckRow]) -> None:
    missing = [path for path in REQUIRED_FILES if not (ROOT / path).exists()]
    add_check(
        rows,
        "FAIL" if missing else "PASS",
        "files",
        "required task files",
        f"missing={missing}" if missing else f"all {len(REQUIRED_FILES)} required files exist",
    )


def check_config(rows: list[CheckRow]) -> dict[str, Any]:
    path = ROOT / "config" / "task_design_plan2.json"
    config = read_json(path)
    expectations = {
        "learning.blocks": (config["learning"]["blocks"], EXPECTED_BLOCKS),
        "learning.learningTrialsPerBlock": (
            config["learning"]["learningTrialsPerBlock"],
            EXPECTED_LEARNING_TRIALS_PER_BLOCK,
        ),
        "test.testTrialsPerBlock": (
            config["test"]["testTrialsPerBlock"],
            EXPECTED_TEST_TRIALS_PER_BLOCK,
        ),
        "test.testOptions": (config["test"]["testOptions"], 5),
        "test.mainTaskReplayAllowed": (config["test"].get("mainTaskReplayAllowed"), False),
    }
    mismatches = [
        f"{key}={observed} expected {expected}"
        for key, (observed, expected) in expectations.items()
        if observed != expected
    ]
    add_check(
        rows,
        "FAIL" if mismatches else "PASS",
        "config",
        "Plan 2 constants",
        "; ".join(mismatches) if mismatches else "Plan 2 constants match the adopted design",
    )
    return config


def check_stimulus_lists(rows: list[CheckRow]) -> dict[str, list[dict[str, Any]]]:
    payload = read_json(ROOT / "stimuli" / "participant_lists_plan2.json")
    lists = payload.get("lists", {})
    list_ids = sorted(lists)
    if list_ids != ["A", "B", "C", "D"]:
        add_check(rows, "FAIL", "stimuli", "counterbalanced lists", f"found list ids={list_ids}")
    else:
        add_check(rows, "PASS", "stimuli", "counterbalanced lists", "found A/B/C/D")

    failures = []
    for list_id, list_rows in sorted(lists.items()):
        if len(list_rows) != EXPECTED_WORDS_PER_LIST:
            failures.append(f"{list_id}: n={len(list_rows)}")
            continue
        control = sum(1 for row in list_rows if row.get("contrast") == "control")
        hard = len(list_rows) - control
        if control != 10 or hard != 10:
            failures.append(f"{list_id}: control={control}, hard={hard}")
        contrast_counts: dict[str, int] = {}
        for row in list_rows:
            contrast_counts[str(row.get("contrast"))] = contrast_counts.get(str(row.get("contrast")), 0) + 1
        for contrast in ["r_l", "v_b", "theta_s", "cluster_r_l", "cluster_v_b"]:
            if contrast_counts.get(contrast) != 2:
                failures.append(f"{list_id}: {contrast}={contrast_counts.get(contrast, 0)}")
    add_check(
        rows,
        "FAIL" if failures else "PASS",
        "stimuli",
        "list composition",
        "; ".join(failures) if failures else "each list has 10 control and 10 hard words with 2 per hard contrast",
    )
    return lists


def check_audio(rows: list[CheckRow]) -> None:
    manifest_path = ROOT / "audio" / "manifest.json"
    manifest = read_json(manifest_path)
    items = manifest.get("items", [])
    missing = [
        item["filename"]
        for item in items
        if not (ROOT / "audio" / item["filename"]).exists()
    ]
    main_count = sum(1 for item in items if item.get("kind") == "main")
    practice_count = sum(1 for item in items if item.get("kind") == "practice")
    if missing:
        add_check(rows, "FAIL", "audio", "audio files", f"missing={missing}")
    elif main_count != EXPECTED_MAIN_AUDIO or practice_count != EXPECTED_PRACTICE_AUDIO:
        add_check(
            rows,
            "FAIL",
            "audio",
            "audio manifest counts",
            f"main={main_count}, practice={practice_count}",
        )
    else:
        add_check(rows, "PASS", "audio", "audio files", "45 MP3 files exist: 40 main and 5 practice")

    review_path = DEFAULT_OUT_DIR / "audio_review_template.csv"
    review_rows = read_csv_rows(review_path)
    if not review_rows:
        add_check(rows, "WARN", "audio", "manual audio review", "audio_review_template.csv not found")
        return

    bad_audibility = [row["word"] for row in review_rows if falsey(row.get("audibleInChrome"))]
    bad_pronunciation = [row["word"] for row in review_rows if falsey(row.get("pronunciationAcceptable"))]
    confused_targets = [row["word"] for row in review_rows if truthy(row.get("soundsLikeOtherTarget"))]
    english_like = [row["word"] for row in review_rows if truthy(row.get("soundsLikeEnglishWord"))]
    blank_review = [
        row["word"]
        for row in review_rows
        if row.get("audibleInChrome", "") == "" or row.get("pronunciationAcceptable", "") == ""
    ]

    if bad_audibility or bad_pronunciation or confused_targets:
        add_check(
            rows,
            "FAIL",
            "audio",
            "manual audio review",
            f"bad_audibility={bad_audibility}; bad_pronunciation={bad_pronunciation}; target_confusions={confused_targets}",
        )
    elif blank_review:
        add_check(
            rows,
            "WARN",
            "audio",
            "manual audio review",
            f"{len(blank_review)} rows still need audible/pronunciation review",
        )
    else:
        add_check(rows, "PASS", "audio", "manual audio review", "all audio review rows are marked acceptable")

    if english_like:
        add_check(rows, "WARN", "audio", "English-like words", f"review flagged={english_like}")


def check_images(rows: list[CheckRow]) -> None:
    image_paths = sorted((ROOT / "images" / "objects").glob("*.svg"))
    main_images = [path for path in image_paths if path.name.startswith("object_")]
    practice_images = [path for path in image_paths if path.name.startswith("practice_")]
    if len(main_images) != EXPECTED_MAIN_IMAGES or len(practice_images) != EXPECTED_PRACTICE_IMAGES:
        add_check(
            rows,
            "FAIL",
            "images",
            "image counts",
            f"main={len(main_images)}, practice={len(practice_images)}",
        )
    else:
        add_check(rows, "PASS", "images", "image counts", "40 main SVGs and 5 practice SVGs exist")

    summary_path = DEFAULT_OUT_DIR / "image_similarity_summary.csv"
    summary_rows = read_csv_rows(summary_path)
    if not summary_rows:
        add_check(rows, "WARN", "images", "image QA output", "image_similarity_summary.csv not found")
        return

    summary = {row["metric"]: parse_float(row["value"]) for row in summary_rows}
    flagged_option_sets = int(summary.get("flaggedTestOptionSets") or 0)
    checked = int(summary.get("testOptionSetsChecked") or 0)
    add_check(
        rows,
        "FAIL" if flagged_option_sets else "PASS",
        "images",
        "5AFC image similarity",
        f"flagged option sets={flagged_option_sets}/{checked}",
    )


def participant_ids(count: int) -> list[str]:
    return [f"QA{index:04d}" for index in range(1, count + 1)]


def check_schedules(rows: list[CheckRow], participants: int) -> None:
    failures: list[str] = []
    list_counts: dict[str, int] = {}
    for participant_id in participant_ids(participants):
        schedule = build_schedule(participant_id)
        list_id = schedule["listId"]
        list_counts[list_id] = list_counts.get(list_id, 0) + 1
        if len(schedule["learningBlocks"]) != EXPECTED_BLOCKS:
            failures.append(f"{participant_id}: learningBlocks={len(schedule['learningBlocks'])}")
        if len(schedule["testBlocks"]) != EXPECTED_BLOCKS:
            failures.append(f"{participant_id}: testBlocks={len(schedule['testBlocks'])}")

        for block in schedule["learningBlocks"]:
            if len(block["trials"]) != EXPECTED_LEARNING_TRIALS_PER_BLOCK:
                failures.append(f"{participant_id} block {block['block']}: learning trials={len(block['trials'])}")
            event_counts: dict[int, int] = {}
            for trial in block["trials"]:
                if len(trial["pairIds"]) != 3 or len(set(trial["pairIds"])) != 3:
                    failures.append(f"{participant_id} block {block['block']} trial {trial['blockTrial']}: invalid pairIds")
                if len(trial["objectIds"]) != 3 or len(set(trial["objectIds"])) != 3:
                    failures.append(f"{participant_id} block {block['block']} trial {trial['blockTrial']}: invalid objectIds")
                visual_families = [object_visual_family(int(object_id)) for object_id in trial["objectIds"]]
                if len(visual_families) != len(set(visual_families)):
                    failures.append(f"{participant_id} block {block['block']} trial {trial['blockTrial']}: visual-family collision")
                for pair_id in trial["wordOrderPairIds"]:
                    event_counts[int(pair_id)] = event_counts.get(int(pair_id), 0) + 1
            if set(event_counts.values()) != {3} or len(event_counts) != EXPECTED_WORDS_PER_LIST:
                failures.append(f"{participant_id} block {block['block']}: unbalanced learning encounters")

        for block in schedule["testBlocks"]:
            if len(block["trials"]) != EXPECTED_TEST_TRIALS_PER_BLOCK:
                failures.append(f"{participant_id} block {block['block']}: test trials={len(block['trials'])}")
            target_positions: dict[int, int] = {}
            target_counts: dict[int, int] = {}
            for trial in block["trials"]:
                option_pair_ids = [int(value) for value in trial["optionPairIds"]]
                option_object_ids = [int(value) for value in trial["optionObjectIds"]]
                if len(option_pair_ids) != 5 or len(set(option_pair_ids)) != 5:
                    failures.append(f"{participant_id} block {block['block']} trial {trial['blockTrial']}: invalid 5AFC pair options")
                if len(option_object_ids) != 5 or len(set(option_object_ids)) != 5:
                    failures.append(f"{participant_id} block {block['block']} trial {trial['blockTrial']}: invalid 5AFC object options")
                if int(trial["targetPairId"]) not in option_pair_ids:
                    failures.append(f"{participant_id} block {block['block']} trial {trial['blockTrial']}: missing target pair")
                if int(trial["targetObjectId"]) not in option_object_ids:
                    failures.append(f"{participant_id} block {block['block']} trial {trial['blockTrial']}: missing target object")
                target_positions[int(trial["targetPosition"])] = target_positions.get(int(trial["targetPosition"]), 0) + 1
                target_counts[int(trial["targetPairId"])] = target_counts.get(int(trial["targetPairId"]), 0) + 1
            if target_positions != {1: 4, 2: 4, 3: 4, 4: 4, 5: 4}:
                failures.append(f"{participant_id} block {block['block']}: target positions={target_positions}")
            if set(target_counts.values()) != {1} or len(target_counts) != EXPECTED_WORDS_PER_LIST:
                failures.append(f"{participant_id} block {block['block']}: target counts invalid")

    add_check(
        rows,
        "FAIL" if failures else "PASS",
        "schedule",
        f"{participants} deterministic schedules",
        f"{len(failures)} failures; first={failures[:5]}" if failures else "learning/test schedule invariants passed",
    )
    add_check(rows, "INFO", "schedule", "list assignment distribution", json.dumps(list_counts, sort_keys=True))


def check_simulation_outputs(rows: list[CheckRow], posterior_threshold: float) -> None:
    scenario_root = ROOT / "analysis" / "simulation_outputs" / "scenario_sweep"
    if not scenario_root.exists():
        add_check(rows, "WARN", "simulation", "scenario sweep", "scenario_sweep output not found")
        return

    criteria = {
        "strong_signal": (0.75, "FAIL"),
        "balanced": (0.60, "WARN"),
        "late_switch": (0.55, "WARN"),
        "weak_signal": (0.50, "WARN"),
    }
    for scenario, (minimum_within_two, bad_status) in criteria.items():
        path = scenario_root / scenario / "switch_recovery_summary.csv"
        summary_rows = read_csv_rows(path)
        if not summary_rows:
            add_check(rows, "WARN", "simulation", f"{scenario} HMM recovery", f"{path} not found")
            continue
        all_rows = [
            row for row in summary_rows
            if row.get("scope") == "all"
            and abs((parse_float(row.get("posteriorThreshold")) or -1) - posterior_threshold) < 1e-9
        ]
        if not all_rows:
            add_check(rows, "WARN", "simulation", f"{scenario} HMM recovery", f"threshold {posterior_threshold} not found")
            continue
        row = all_rows[0]
        within_two = parse_float(row.get("withinTwoEncounterRecoveryRate")) or 0.0
        state_accuracy = parse_float(row.get("stateAccuracyEventLevel"))
        status = "PASS" if within_two >= minimum_within_two else bad_status
        add_check(
            rows,
            status,
            "simulation",
            f"{scenario} HMM recovery",
            f"withinTwo={within_two:.3f}, stateAccuracy={state_accuracy}, criterion>={minimum_within_two:.2f}",
        )

    benchmark_path = ROOT / "analysis" / "simulation_outputs" / "method_benchmark" / "method_recovery_summary.csv"
    benchmark_rows = read_csv_rows(benchmark_path)
    if not benchmark_rows:
        add_check(rows, "WARN", "simulation", "method benchmark", "method_recovery_summary.csv not found")
        return
    for scenario in ["balanced", "weak_signal"]:
        scenario_rows = [
            row for row in benchmark_rows
            if row.get("simulationScenario") == scenario and row.get("scope") == "all"
        ]
        if not scenario_rows:
            add_check(rows, "WARN", "simulation", f"{scenario} method benchmark", "no rows found")
            continue
        best = max(
            scenario_rows,
            key=lambda row: parse_float(row.get("withinTwoEncounterRecoveryRate")) or -1.0,
        )
        online_rows = [row for row in scenario_rows if parse_int(row.get("usesFutureData")) == 0]
        best_online = max(
            online_rows,
            key=lambda row: parse_float(row.get("withinTwoEncounterRecoveryRate")) or -1.0,
        ) if online_rows else None
        detail = (
            f"best={best['method']} withinTwo={parse_float(best.get('withinTwoEncounterRecoveryRate')):.3f}"
        )
        if best_online:
            detail += (
                f"; bestOnline={best_online['method']} "
                f"withinTwo={parse_float(best_online.get('withinTwoEncounterRecoveryRate')):.3f}"
            )
        add_check(rows, "INFO", "simulation", f"{scenario} method benchmark", detail)


def require_openpyxl() -> Any:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise SystemExit("openpyxl is required for workbook validation.") from error
    return load_workbook


def sheet_rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = [str(value) for value in values[0]]
    return [dict(zip(headers, row)) for row in values[1:]]


def check_workbook(rows: list[CheckRow], workbook_path: Path) -> None:
    load_workbook = require_openpyxl()
    if not workbook_path.exists():
        add_check(rows, "FAIL", "workbook", workbook_path.name, "file not found")
        return
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    missing_sheets = [name for name in REQUIRED_WORKBOOK_SHEETS if name not in workbook.sheetnames]
    add_check(
        rows,
        "FAIL" if missing_sheets else "PASS",
        "workbook",
        f"{workbook_path.name} sheets",
        f"missing={missing_sheets}" if missing_sheets else "all required sheets found",
    )
    if missing_sheets:
        return

    model_ready = sheet_rows(workbook, "ModelReady")
    model_headers = list(model_ready[0].keys()) if model_ready else []
    missing_columns = [name for name in REQUIRED_MODEL_READY_COLUMNS if name not in model_headers]
    add_check(
        rows,
        "FAIL" if missing_columns else "PASS",
        "workbook",
        f"{workbook_path.name} ModelReady columns",
        f"missing={missing_columns}" if missing_columns else "all required ModelReady columns found",
    )

    counts = {
        "LearningEvents": len(sheet_rows(workbook, "LearningEvents")),
        "LearningTrials": len(sheet_rows(workbook, "LearningTrials")),
        "Data": len(sheet_rows(workbook, "Data")),
        "ModelReady": len(model_ready),
    }
    expected_counts = {
        "LearningEvents": EXPECTED_LEARNING_EVENTS,
        "LearningTrials": EXPECTED_LEARNING_TRIALS,
        "Data": EXPECTED_TEST_TRIALS,
        "ModelReady": EXPECTED_MODEL_READY_ROWS,
    }
    count_failures = [
        f"{sheet}={observed} expected {expected_counts[sheet]}"
        for sheet, observed in counts.items()
        if observed != expected_counts[sheet]
    ]
    add_check(
        rows,
        "FAIL" if count_failures else "PASS",
        "workbook",
        f"{workbook_path.name} row counts",
        "; ".join(count_failures) if count_failures else json.dumps(counts, sort_keys=True),
    )

    observation_seq = [parse_int(row.get("observationSeq")) for row in model_ready]
    expected_seq = list(range(1, len(model_ready) + 1))
    add_check(
        rows,
        "FAIL" if observation_seq != expected_seq else "PASS",
        "workbook",
        f"{workbook_path.name} chronological ModelReady",
        "observationSeq is not consecutive" if observation_seq != expected_seq else "observationSeq is consecutive",
    )

    test_schedule = sheet_rows(workbook, "TestSchedule")
    target_position_failures = []
    for block in range(1, EXPECTED_BLOCKS + 1):
        block_rows = [row for row in test_schedule if parse_int(row.get("block")) == block]
        counts_by_position: dict[int, int] = {}
        for row in block_rows:
            position = parse_int(row.get("targetPosition"))
            if position is not None:
                counts_by_position[position] = counts_by_position.get(position, 0) + 1
        if counts_by_position != {1: 4, 2: 4, 3: 4, 4: 4, 5: 4}:
            target_position_failures.append(f"block {block}: {counts_by_position}")
    add_check(
        rows,
        "FAIL" if target_position_failures else "PASS",
        "workbook",
        f"{workbook_path.name} target-position balance",
        "; ".join(target_position_failures) if target_position_failures else "each 5AFC position is correct four times per block",
    )

    data_rows = sheet_rows(workbook, "Data")
    timeout_failures = []
    for row in data_rows:
        timed_out = truthy(row.get("responseTimedOut")) or truthy(row.get("timedOut"))
        if timed_out and not truthy(row.get("noResponse")):
            timeout_failures.append(f"block {row.get('block')} trial {row.get('blockTrial')}: noResponse not set")
        if timed_out and truthy(row.get("correct")):
            timeout_failures.append(f"block {row.get('block')} trial {row.get('blockTrial')}: timeout marked correct")
    add_check(
        rows,
        "FAIL" if timeout_failures else "PASS",
        "workbook",
        f"{workbook_path.name} timeout coding",
        "; ".join(timeout_failures[:5]) if timeout_failures else "timeout coding is internally consistent",
    )


def markdown_report(rows: list[CheckRow]) -> str:
    status = overall_status(rows)
    counts: dict[str, int] = {}
    for row in rows:
        counts[row.status] = counts.get(row.status, 0) + 1

    lines = [
        "# Pre-Pilot Readiness Report",
        "",
        f"Overall status: **{status}**",
        "",
        "## Status Counts",
        "",
        "| Status | Count |",
        "| --- | ---: |",
    ]
    for status_name in ["FAIL", "WARN", "PASS", "INFO"]:
        lines.append(f"| {status_name} | {counts.get(status_name, 0)} |")

    lines.extend([
        "",
        "## Checks",
        "",
        "| Status | Area | Check | Detail |",
        "| --- | --- | --- | --- |",
    ])
    for row in sorted(rows, key=lambda item: (-status_rank(item.status), item.area, item.check)):
        lines.append(
            f"| {row.status} | {row.area} | {row.check} | {escape_markdown(row.detail)} |"
        )

    lines.extend([
        "",
        "## Interpretation",
        "",
        "- `FAIL`: fix before pilot collection or before trusting the workbook.",
        "- `WARN`: acceptable for internal dry runs, but resolve or document before formal pilot collection.",
        "- `INFO`: diagnostic context that should not block collection by itself.",
        "",
    ])
    return "\n".join(lines)


def escape_markdown(text: str) -> str:
    return str(text).replace("|", "\\|").replace("\n", " ")


def write_report(path: Path, rows: list[CheckRow]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(markdown_report(rows), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--participants", type=int, default=80)
    parser.add_argument("--posterior-threshold", type=float, default=0.70)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument("--refresh", action="store_true", help="Run QA and simulation scripts before reporting.")
    parser.add_argument("--skip-simulation", action="store_true", help="When used with --refresh, skip simulation refresh.")
    parser.add_argument("--workbook", type=Path, action="append", default=[], help="Optional exported .xlsx workbook to validate.")
    args = parser.parse_args()
    if args.participants <= 0:
        raise SystemExit("--participants must be positive")
    if not 0 < args.posterior_threshold < 1:
        raise SystemExit("--posterior-threshold must be between 0 and 1")
    return args


def main() -> None:
    args = parse_args()
    rows: list[CheckRow] = []

    if args.refresh:
        run_refresh(args, rows)

    check_required_files(rows)
    check_config(rows)
    check_stimulus_lists(rows)
    check_audio(rows)
    check_images(rows)
    check_schedules(rows, args.participants)
    check_simulation_outputs(rows, args.posterior_threshold)
    for workbook_path in args.workbook:
        check_workbook(rows, workbook_path)

    csv_path = args.out_dir / "prepilot_readiness_checks.csv"
    report_path = args.out_dir / "prepilot_readiness_report.md"
    write_csv(csv_path, rows)
    write_report(report_path, rows)
    status = overall_status(rows)
    print(f"Overall status: {status}")
    print(f"Wrote {csv_path}")
    print(f"Wrote {report_path}")
    if status == "FAIL":
        raise SystemExit(1)


if __name__ == "__main__":
    main()
