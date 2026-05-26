#!/usr/bin/env python3
"""
Summarize a CSSL Validation workbook exported by the browser task.

The script reads the `ModelReady` sheet and writes compact CSV summaries for
pilot checks. It intentionally starts with transparent descriptive analyses
before any HMM or mixture modeling.
"""

from __future__ import annotations

import argparse
import csv
import statistics
from collections import defaultdict
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUT_DIR = ROOT / "analysis" / "participant_summaries"


def require_openpyxl() -> Any:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise SystemExit("openpyxl is required to read .xlsx exports.") from error
    return load_workbook


def read_model_ready(path: Path) -> list[dict[str, Any]]:
    load_workbook = require_openpyxl()
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "ModelReady" not in workbook.sheetnames:
        raise SystemExit(f"{path} does not contain a ModelReady sheet")
    sheet = workbook["ModelReady"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value) for value in rows[0]]
    records = []
    for values in rows[1:]:
        record = {header: cell for header, cell in zip(headers, values)}
        records.append(record)
    return records


def value_to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    return int(float(value))


def value_to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    return float(value)


def mean(values: list[float | int]) -> float | str:
    return round(statistics.fmean(values), 6) if values else ""


def rate(rows: list[dict[str, Any]], key: str) -> float | str:
    values = [value_to_int(row.get(key)) for row in rows]
    clean = [value for value in values if value is not None]
    return mean(clean)


def accuracy(rows: list[dict[str, Any]]) -> float | str:
    return rate(rows, "correct")


def participant_value(rows: list[dict[str, Any]], key: str) -> Any:
    for row in rows:
        if row.get(key) not in (None, ""):
            return row.get(key)
    return ""


def hard_bucket(row: dict[str, Any]) -> str:
    return "control" if row.get("contrast") == "control" else "hard"


def build_overall_summary(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    learning_rows = [row for row in rows if row.get("observationType") == "learning_3afc"]
    test_rows = [row for row in rows if row.get("observationType") == "test_5afc"]
    previous_correct_rows = [
        row for row in learning_rows
        if value_to_int(row.get("previousCorrect")) == 1
    ]
    previous_incorrect_rows = [
        row for row in learning_rows
        if value_to_int(row.get("previousIncorrect")) == 1
    ]
    p_after_correct = accuracy(previous_correct_rows)
    p_after_incorrect = accuracy(previous_incorrect_rows)
    contingency_delta = (
        round(float(p_after_correct) - float(p_after_incorrect), 6)
        if p_after_correct != "" and p_after_incorrect != "" else ""
    )

    return [{
        "participantId": participant_value(rows, "participantId"),
        "listId": participant_value(rows, "listId"),
        "modelReadyRows": len(rows),
        "learningRows": len(learning_rows),
        "testRows": len(test_rows),
        "learningAccuracy": accuracy(learning_rows),
        "testAccuracy": accuracy(test_rows),
        "testAccuracyControl": accuracy([row for row in test_rows if hard_bucket(row) == "control"]),
        "testAccuracyHard": accuracy([row for row in test_rows if hard_bucket(row) == "hard"]),
        "testNoResponseRate": rate(test_rows, "noResponse"),
        "testTimeoutRate": rate(test_rows, "timedOut"),
        "pLearningCorrectAfterPreviousCorrect": p_after_correct,
        "pLearningCorrectAfterPreviousIncorrect": p_after_incorrect,
        "learningContingencyDelta": contingency_delta,
    }]


def build_block_summary(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[Any, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[(row.get("block"), str(row.get("observationType")), hard_bucket(row))].append(row)

    summary = []
    for (block, observation_type, difficulty), group_rows in sorted(groups.items()):
        summary.append({
            "block": block,
            "observationType": observation_type,
            "difficulty": difficulty,
            "n": len(group_rows),
            "accuracy": accuracy(group_rows),
            "noResponseRate": rate(group_rows, "noResponse"),
            "timeoutRate": rate(group_rows, "timedOut"),
            "meanRtMs": mean([
                value for value in (value_to_float(row.get("rtMs")) for row in group_rows)
                if value is not None
            ]),
        })
    return summary


def build_word_trajectory(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[Any, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[row.get("pairId")].append(row)

    trajectories = []
    for pair_id, group_rows in sorted(groups.items(), key=lambda item: int(item[0])):
        group_rows.sort(key=lambda row: (int(row.get("block") or 0), int(row.get("observationSeq") or 0)))
        first = group_rows[0]
        learning = [row for row in group_rows if row.get("observationType") == "learning_3afc"]
        tests = [row for row in group_rows if row.get("observationType") == "test_5afc"]
        learning_by_encounter = sorted(
            learning,
            key=lambda row: int(row.get("encounterIndex") or 0),
        )
        trajectories.append({
            "pairId": pair_id,
            "word": first.get("word"),
            "contrast": first.get("contrast"),
            "phonology": first.get("phonology"),
            "learningAccuracy": accuracy(learning),
            "testAccuracy": accuracy(tests),
            "learningCorrectSequence": "".join(
                str(value_to_int(row.get("correct")) or 0) for row in learning_by_encounter
            ),
            "testCorrectByBlock": "|".join(
                f"{row.get('block')}:{value_to_int(row.get('correct')) or 0}" for row in tests
            ),
            "timeoutByBlock": "|".join(
                f"{row.get('block')}:{value_to_int(row.get('timedOut')) or 0}" for row in tests
            ),
        })
    return trajectories


def build_timeout_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output = []
    for row in rows:
        if value_to_int(row.get("noResponse")) == 1 or value_to_int(row.get("timedOut")) == 1:
            output.append({
                "participantId": row.get("participantId"),
                "observationSeq": row.get("observationSeq"),
                "observationType": row.get("observationType"),
                "block": row.get("block"),
                "pairId": row.get("pairId"),
                "word": row.get("word"),
                "contrast": row.get("contrast"),
                "noResponse": value_to_int(row.get("noResponse")),
                "timedOut": value_to_int(row.get("timedOut")),
                "rtMs": row.get("rtMs"),
            })
    return output


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows and not fieldnames:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = fieldnames or list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def safe_stem(path: Path) -> str:
    return "".join(char if char.isalnum() or char in "-_" else "_" for char in path.stem)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rows = read_model_ready(args.workbook)
    if not rows:
        raise SystemExit("ModelReady sheet is empty")
    out_dir = args.out_dir / safe_stem(args.workbook)
    write_csv(out_dir / "overall_summary.csv", build_overall_summary(rows))
    write_csv(out_dir / "block_summary.csv", build_block_summary(rows))
    write_csv(out_dir / "word_trajectory.csv", build_word_trajectory(rows))
    write_csv(out_dir / "timeout_rows.csv", build_timeout_rows(rows), [
        "participantId",
        "observationSeq",
        "observationType",
        "block",
        "pairId",
        "word",
        "contrast",
        "noResponse",
        "timedOut",
        "rtMs",
    ])
    print(f"Wrote summaries to {out_dir}")


if __name__ == "__main__":
    main()
